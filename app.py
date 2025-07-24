from flask import Flask, render_template, request, redirect, flash
import openpyxl
import os
import re

app = Flask(__name__)
app.secret_key = 'clave_secreta_segura'

archivo = "base_socios.xlsx"

# Crear archivo si no existe
if not os.path.exists(archivo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios"
    ws.append([
        "Nombres y Apellidos",
        "Cédula",
        "Parroquia",
        "Comunidad / Dirección",
        "Fecha de Nacimiento",
        "¿Tiene Discapacidad?",
        "¿Tiene Nichos?",
        "Celular",
        "Correo (Opcional)"
    ])
    wb.save(archivo)

# Verificar si cédula ya fue registrada
def cedula_existente(cedula):
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if str(fila[1]).strip() == cedula.strip():
            return True
    return False

@app.route('/', methods=['GET', 'POST'])
def formulario():
    if request.method == 'POST':
        datos = {
            "nombre": request.form.get('nombre', '').strip(),
            "cedula": request.form.get('cedula', '').strip(),
            "parroquia": request.form.get('parroquia', '').strip(),
            "direccion": request.form.get('direccion', '').strip(),
            "fecha_nacimiento": request.form.get('fecha_nacimiento', '').strip(),
            "discapacidad": request.form.get('discapacidad', '').strip(),
            "nichos": request.form.get('nichos', '').strip(),
            "celular": request.form.get('celular', '').strip(),
            "correo": request.form.get('correo', '').strip()
        }

        # Validar campos obligatorios (todos excepto correo)
        campos_obligatorios = ["nombre", "cedula", "parroquia", "direccion",
                               "fecha_nacimiento", "discapacidad", "nichos", "celular"]
        for campo in campos_obligatorios:
            if datos[campo] == '':
                flash(f"Por favor completa el campo obligatorio: {campo.replace('_', ' ').title()}.", "error")
                return render_template('formulario.html', datos=datos, error_campo=campo)

        # Validar formato de fecha DD/MM/AAAA
        patron_fecha = r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/[0-9]{4}$"
        if not re.match(patron_fecha, datos["fecha_nacimiento"]):
            flash("La fecha de nacimiento debe tener formato DD/MM/AAAA.", "error")
            datos["fecha_nacimiento"] = ''
            return render_template('formulario.html', datos=datos, error_campo="fecha_nacimiento")

        # Verificar cédula duplicada
        if cedula_existente(datos["cedula"]):
            flash(f"La cédula {datos['cedula']} ya fue registrada.", "error")
            datos["cedula"] = ''
            return render_template('formulario.html', datos=datos, error_campo="cedula")

        # Guardar datos en Excel
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        ws.append([
            datos["nombre"],
            datos["cedula"],
            datos["parroquia"],
            datos["direccion"],
            datos["fecha_nacimiento"],
            datos["discapacidad"],
            datos["nichos"],
            datos["celular"],
            datos["correo"]
        ])
        wb.save(archivo)

        flash("✅ Datos guardados exitosamente.", "success")
        return redirect('/')

    # GET: mostrar formulario vacío
    return render_template('formulario.html', datos={}, error_campo=None)


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))  # Render asigna el puerto dinámicamente
    app.run(host='0.0.0.0', port=port, debug=True)
